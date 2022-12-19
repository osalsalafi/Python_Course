# -------------------------------------------------------------------------------------------------
# -- Python Course => Dealing with Excel spreadsheets => 02. Open an Excel file and read from it --
# -------------------------------------------------------------------------------------------------
import openpyxl
from pathlib import Path

excelfile = openpyxl.load_workbook('C:\\My Git-Hub\\Python_Course\\Hsoub Academy\\Python_Course 2\\Dealing with Excel spreadsheets\\test1.xlsx')
print(excelfile.sheetnames) # ['Sheet1', 'Sheet2', 'Sheet3']

Sheet1 = excelfile['Sheet1']
print(Sheet1.title) # Sheet1

activeSheet = excelfile.active
print(activeSheet.title) # Sheet1

# to print the items of the cells
print(Sheet1['A1'].value) # type: ignore # osama
print(Sheet1['B1'].value) # type: ignore # 1000
print(Sheet1['C1'].row) # type: ignore # 1
print(Sheet1['C1'].column) # type: ignore # 3
print(Sheet1['C1'].coordinate) # type: ignore # C1
print(Sheet1.cell(row=1,column=3).value) # type: ignore # 2010-10-01 00:00:00

print(Sheet1.max_column) # type: ignore # 3
print(Sheet1.max_row) # type: ignore # 6

