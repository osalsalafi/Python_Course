# --------------------------------------------------------------------------------------------------
# -- Python Course => Dealing with Excel spreadsheets => 03. Create an Excel file and write to it --
# --------------------------------------------------------------------------------------------------
import openpyxl
from pathlib import Path
import os
# os.chdir('C:\\My Git-Hub\\Python_Course\\Hsoub Academy\\Python_Course 2\\Dealing with Excel spreadsheets')

# create excel file
excelfile = openpyxl.Workbook()
print(excelfile.sheetnames)

# change sheet name
sheetname = excelfile.active
sheetname.title = 'ورقة 1'

# create sheet
excelfile.create_sheet('ورقة 2')
excelfile.create_sheet(index=2,  title='ورقة 3')  # type: ignore

# delete sheet
del excelfile['ورقة 3']

# write to excel sheet
sheet = excelfile['ورقة 1']
sheet['A1'] = ''  # type: ignore

# try to write 4 names in the 3rd column in the first sheet
names = ['Osama', 'Salem', 'Ali', 'Wafaa']
salaries = [1000, 1500, 2000, 2500]
for index, name in enumerate(names) :
    sheet[f'C{index+1}'].value = name  # type: ignore

for index, salary in enumerate(salaries) :
    sheet.cell(row=index+1, column=4).value = salary  # type: ignore

# save excel file
excelfile.save('C:\\My Git-Hub\\Python_Course\\Hsoub Academy\\Python_Course 2\\Dealing with Excel spreadsheets\\new excel.xlsx')